from pprint import pprint
import random as rn
import math
import copy
from random import randrange
import matplotlib.pyplot as plt
import time
import openpyxl
import os

import numpy as np
from collections import defaultdict

Rosenbrock = "Rosenbrock"
Step = "Step"
Ackley = "Ackley"
Griewank = "Griewank"
Rastrigin = "Rastrigin"
Genr_Penalized = "Generalized Penalized"

FUNCTIONS = [(Rosenbrock, 30), (Step, 100), (Ackley, 32),
             (Griewank, 600), (Rastrigin, 512), (Genr_Penalized, 50)]

TEST_FUNC = None
BOUND = None
# Ns = [10, 30, 50]
Ns = [50]

wMax = 0.9
wMin = 0.2
wC = 0.7


# 40000 evaluation
ITERATIONS = 500
SWARM_SIZE = 80

DEBUG = False
TIMER_MODE = False
EXEl_WRITE = False

c1Min = c2Min = 0.5
c1Max = c2Max = 2.5


class Particle(object):

    w = wMax
    ind = 0
    swarm_size = None

    def __init__(self, n, size):
        self.n = n
        self.swarm_size = size
        self.x = [-1] * n
        self.v = [-1] * n
        self.f = None
        self.pbest = None
        self.pbestVal = None
        self.positioning()

    def positioning(self):

        # k = BOUND/self.swarm_size

        # for i in range(self.n):
        #     domain = BOUND - (k * self.ind)

        #     self.x[i] = rn.uniform(-domain, domain)
        #     self.v[i] = 0.1 * self.x[i]

        # self.ind += 1

        for i in range(self.n):
            self.x[i] = rn.uniform(-BOUND, BOUND)
            self.v[i] = 0.1 * self.x[i]

        self.calculate_f()
        self.pbest = self.x
        self.pbestVal = self.f

    def calculate_f(self):
        fVal = 0
        if TEST_FUNC == Rosenbrock:
            for i in range(self.n-1):
                fVal += (100 * (self.x[i + 1] - self.x[i]
                                ** 2) ** 2) + ((self.x[i] - 1) ** 2)

        if TEST_FUNC == Step:
            for i in range(self.n):
                fVal += (self.x[i] + 0.5) ** 2

        elif TEST_FUNC == Ackley:
            part1 = part2 = 0
            for i in range(self.n):
                part1 += self.x[i] ** 2
                part2 += math.cos(2 * math.pi * self.x[i])

            fVal = -20 * math.exp(-0.2 * math.sqrt((1/self.n) * part1)) - \
                math.exp((1/self.n) * part2) + 20 + math.e

        elif TEST_FUNC == Griewank:
            for i in range(self.n):
                fVal += self.x[i] ** 2

            fVal *= 1/4000

            temp = 1
            for i in range(self.n):
                temp *= math.cos(self.x[i] / math.sqrt(i+1))

            fVal = fVal - temp + 1

        elif TEST_FUNC == Rastrigin:
            for i in range(self.n):
                fVal += self.x[i]**2 - 10 * \
                    math.cos(2 * math.pi*self.x[i]) + 10

        elif TEST_FUNC == Genr_Penalized:
            fVal = self.gen_penalized()

        self.f = fVal

    def gen_penalized(self):
        fVal = 0
        pi = math.pi
        sigma1 = sigma2 = 0
        for i in range(self.n-1):
            yi = 1 + 1/4 * (self.x[i]+1)
            yi1 = 1 + 1/4 * (self.x[i+1]+1)

            sigma1 += ((yi - 1) ** 2) * (1 + 10 * math.sin(pi * yi1) ** 2)

        for i in range(self.n):
            sigma2 += self.u(self.x[i], 10, 100, 4)

        y1 = 1 + 1/4 * (self.x[0]+1)
        yn = 1 + 1/4 * (self.x[self.n-1]+1)

        fVal = pi/self.n * (10 * math.sin(pi * y1) +
                            sigma1 + (yn-1) ** 2) + sigma2

        return fVal

    def u(self, xi, a, k, m):
        if xi > a:
            return k * (xi-a)**m
        elif xi >= -a:
            return 0
        else:
            return k * (-xi-a)**m

    def updatePbest(self):
        if self.f < self.pbestVal:
            self.pbest = self.x
            self.pbestVal = self.f

    def updateVelocityAndPos(self, gbest, t):
        for i in range(self.n):
            c1 = rn.uniform(0, 2)
            c2 = 4 - c1

            # c1 = (c1Min - c1Max) * (t/ITERATIONS) + c1Min
            # c2 = (c2Min - c2Max) * (t/ITERATIONS) + c2Min

            r1 = rn.random()
            r2 = rn.random()

            self.v[i] = (self.w * self.v[i]) + (c1 * r1 *
                                                (self.pbest[i] - self.x[i])) + (c2 * r2 * (gbest[i] - self.x[i]))

            self.x[i] += self.v[i]
            self.calculate_f()

        self.updateW(t)

    def updateW(self, t):
        # self.w *= wC
        # self.w = wMax/math.log(t+1)
        self.w = math.exp(-0.9 * t+1)*wMax

        self.w = min(self.w, wMin)

    def __str__(self):
        return "x: " + str(self.x[:5]) + "f(x): " + str(self.f)

    def __repr__(self):
        return str(self)


class Swarm(object):

    def __init__(self, size, n):
        self.size = size
        self.n = n
        self.gbest = None
        self.gbestVal = None
        self.particles = []
        self.initial()

    def initial(self):
        for _ in range(self.size):
            self.particles.append(Particle(self.n, self.size))

        self.gbest = self.particles[0].x
        self.gbestVal = self.particles[0].f

    def updateGbest(self):
        for par in self.particles:
            if par.pbestVal < self.gbestVal:
                self.gbest = par.pbest
                self.gbestVal = par.pbestVal

    def __str__(self):
        return str(self.number) + str(self.getCord())

    def __repr__(self):
        return str(self)


def printResult(answers):

    minAns = min(answers, key=lambda t: t[0])
    avgAns = sum(ans[0] for ans in answers)/len(answers)
    maxAns = max(answers, key=lambda t: t[0])
    variance = round(math.sqrt(np.var([ans[0]for ans in answers])), 3)

    minTime = min(answers, key=lambda t: t[2])
    avgTime = sum(ans[2] for ans in answers)/len(answers)
    maxTime = max(answers, key=lambda t: t[2])

    if TIMER_MODE == False:
        print("\nbest = ", round(minAns[0], 3), "avg = ", round(avgAns, 3),
              " worst = ", round(maxAns[0], 3), "variance = ", round(variance, 3))
        print("----------------")
        print("minTime = ", round(minTime[2], 3), "avgTime = ",
              round(avgTime, 3), " maxTime = ", round(maxTime[2], 3))
    else:
        print("\nbest = ", minAns)


def writeResultToExel(answers, myRow):
    minAns = min(answers, key=lambda t: t[0])
    avgAns = sum(ans[0] for ans in answers)/len(answers)
    maxAns = max(answers, key=lambda t: t[0])
    variance = round(math.sqrt(np.var([ans[0]for ans in answers])), 3)

    minTime = min(answers, key=lambda t: t[2])
    avgTime = sum(ans[2] for ans in answers)/len(answers)
    maxTime = max(answers, key=lambda t: t[2])

    wbkName = 'Results.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    for wks in wbk.worksheets:
        myCol = 2

        wks.cell(row=myRow, column=myCol).value = round(minAns[0], 3)

        if TIMER_MODE == False:
            wks.cell(row=myRow, column=myCol+1).value = round(avgAns, 3)
            wks.cell(row=myRow, column=myCol+2).value = round(maxAns[0], 3)
            wks.cell(row=myRow, column=myCol+3).value = round(variance, 3)

            wks.cell(row=myRow, column=myCol+4).value = round(minTime[2], 3)
            wks.cell(row=myRow, column=myCol+5).value = round(avgTime, 3)
            wks.cell(row=myRow, column=myCol+6).value = round(maxTime[2], 3)

    wbk.save(wbkName)
    wbk.close


def plotResult(generation, bests, averages):

    plt.plot(list(range(generation)), bests, '-', color="gray",
             label='best of generations', linewidth=1.52)

    plt.xlabel('best solution')
    plt.ylabel('generation')

    plt.show()

    plt.plot(list(range(generation)), averages, 'bo-',
             label='avg of generations', linewidth=1.5)

    plt.xlabel('best solution')
    plt.ylabel('generation')

    plt.show()


def plotProgress(bests, avgs):

    plt.plot(list(range(len(bests))), bests, '-', color="gray",
             label='best of generations', linewidth=1.5)

    plt.xlabel('best solution')
    plt.ylabel('generation')

    plt.show()

    plt.plot(list(range(len(avgs))), avgs, 'bo-',
             label='avg of generations', linewidth=1.5)

    plt.xlabel('avg solution')
    plt.ylabel('generation')

    plt.show()


def PSO(SWARM_SIZE, N, ITERATIONS=50, DEBUG=True):

    timer = time.time()
    if TIMER_MODE:
        ITERATIONS = 10000000000

    # particles random initial
    swarm = Swarm(SWARM_SIZE, N)

    for i in range(ITERATIONS):

        for particle in swarm.particles:
            particle.updatePbest()

        swarm.updateGbest()

        for particle in swarm.particles:
            particle.updateVelocityAndPos(swarm.gbest, i+1)

        if DEBUG:
            print("iteration:", i, "gbest: ",
                  swarm.gbestVal, "\t", [round(x, 3) for x in swarm.gbest[:10]])

            # print("iteration:", i, "gbest: ", swarm.gbestVal)

    return (swarm.gbest, swarm.gbestVal)


if __name__ == '__main__':

    row = 1

    if TIMER_MODE:
        run = 1
    else:
        run = 10

    for test_func in FUNCTIONS[4:]:
        (TEST_FUNC, BOUND) = test_func
        row += 3

        for N in Ns:
            print("\n<<<<", TEST_FUNC,
                  "with bound :[", -BOUND, ",", BOUND, "] and N =", N, ">>>>\n")

            answers = []

            for _ in range(run):
                start = time.time()

                (bestX, bestF) = PSO(SWARM_SIZE, N, ITERATIONS, DEBUG)

                duration = round(time.time() - start, 4)

                print('time: ', duration, '\tbest:',
                      round(bestF, 3), '\tx[0:5]:', [round(x, 3) for x in bestX[:5]])

                answers.append(
                    (bestF, bestX, duration))

            printResult(answers)

            if EXEl_WRITE:
                writeResultToExel(answers, row)
            row += 1
