from pprint import pprint
import random as rn
import math
import copy
from random import randrange
import matplotlib.pyplot as plt
import time
import openpyxl
import os
import tsp

from tspy2 import TSP
from tspy2.solvers import NN_solver
from tspy2.solvers import TwoOpt_solver
import numpy as np
from collections import defaultdict

ORDER_2POINT = 'ORDER_2POINT'
ELITISM = "ELITISM"
GENERATIONAL = "GENERATIONAL"
RANDOM = 'RANDOM'

MUTATION_RATE = 0.2
POPULATION_SIZE = 50
MAX_GENERATION = 10
XOVER_METHOD = ORDER_2POINT
SELECTION = RANDOM
SURVIVOR_SEL_TYPE = ELITISM
DEBUG = False

generation = 1
bests = []
averages = []

TIMER_MODE = False
EXEl_WRITE = False


class Node(object):

    def __init__(self, number, x, y):
        self.number = number
        self.x = x
        self.y = y
        self.demand = None
        self.cluster = None

    def getCord(self):
        return (self.x, self.y)

    def __str__(self):
        return str(self.number) + str(self.getCord())

    def __repr__(self):
        return str(self)


class Problem(object):

    def __init__(self):
        self.name = None
        self.dimention = None
        self.capacity = None
        self.depotCluster = 1
        self.clusters = []

    def __str__(self):
        return 'depot cluster: ' + str(self.depotCluster)

    def __repr__(self):
        return str(self)


def printResult(answers):

    minAns = min(answers, key=lambda t: t[1])
    maxAns = max(answers, key=lambda t: t[1])
    variance = round(math.sqrt(np.var([ans[1]for ans in answers])), 3)

    print("\nbest[0:10]=", minAns[0][0:10], "\tmin cost:", minAns[1])
    if TIMER_MODE == False:
        print("worst[0:10]=", maxAns[0][0:10], "\tmax cost:",
              max(answers, key=lambda t: t[1])[1])
        print("\naverage cost:", sum(ans[1] for ans in answers)/len(answers))
        print("\nvariance of costs:", variance)

        print("\nmin time:", min(answers, key=lambda t: t[2])[2])
        print("avg time:", str(sum(float(ans[2])
                                   for ans in answers)/len(answers))[0:6])
        print("max time:", max(answers, key=lambda t: t[2])[2])

    print("\naverage vehicels:", sum(ans[3] for ans in answers)/len(answers))


def writeResultToExel(file_name, answers, myRow):
    minCost = min(answers, key=lambda t: t[1])[1]
    if TIMER_MODE == False:
        maxCost = max(answers, key=lambda t: t[1])[1]
        avgCost = sum(ans[1] for ans in answers)/len(answers)
        costVariance = round(math.sqrt(np.var([ans[1]for ans in answers])), 3)

        minTime = min(answers, key=lambda t: t[2])[2]
        maxTime = max(answers, key=lambda t: t[2])[2]
        avgTime = str(sum(float(ans[2])for ans in answers)/len(answers))[0:6]

    avgVehc = sum(ans[3] for ans in answers)/len(answers)

    wbkName = 'Results.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    for wks in wbk.worksheets:
        myCol = 7

        # wks.cell(row=myRow, column=1).value = file_name

        wks.cell(row=myRow, column=myCol).value = minCost
        if TIMER_MODE == False:
            wks.cell(row=myRow, column=myCol+1).value = avgCost
            wks.cell(row=myRow, column=myCol+2).value = maxCost
            wks.cell(row=myRow, column=myCol+3).value = costVariance

            wks.cell(row=myRow, column=myCol+4).value = minTime
            wks.cell(row=myRow, column=myCol+5).value = avgTime
            wks.cell(row=myRow, column=myCol+6).value = maxTime

        wks.cell(row=myRow, column=myCol+7).value = avgVehc

    wbk.save(wbkName)
    wbk.close


def plotResult(generation, bests, averages):

    plt.plot(list(range(generation)), bests, '-', color="gray",
             label='best of generations', linewidth=1.52)

    plt.xlabel('best solution')
    plt.ylabel('generation')

    plt.show()

    plt.plot(list(range(generation)), averages, 'bo-',
             label='avg of generations', linewidth=1.5)

    plt.xlabel('best solution')
    plt.ylabel('generation')

    plt.show()


def loadInstance(file):
    problem = Problem()
    with open(file) as f:
        lines = [line.rstrip() for line in f]

        problem.name = lines[0].split(':')[1]
        problem.dimention = int(lines[3].split(':')[1])
        problem.capacity = int(lines[4].split(':')[1])

        problem.clusters = {}

        # NODE_COORD_SECTION first index
        cordSecIndex = 8
        # DEMAND_SECTION first index
        demSecIndex = cordSecIndex + problem.dimention + 1
        # CLUSTER_SECTION first index
        cludSecIndex = demSecIndex + problem.dimention + 1

        # Golden type instance . its cluster number start from 1
        if problem.name == '':
            x = -1
        else:
            x = 0

        problem.depotCluster = int(lines[cludSecIndex].split()[1]) + x

        for i in range(problem.dimention):
            data = lines[cordSecIndex+i].split()
            node = Node(int(data[0]), float(data[1]), float(data[2]))

            node.demand = int(lines[demSecIndex+i].split()[1])
            node.cluster = int(lines[cludSecIndex+i].split()[1]) + x
            problem.clusters.setdefault(node.cluster, []).append(node)

    return problem


class Individual(object):

    problem = None
    MList = None

    def __init__(self, chromosome):
        self.chromosome = chromosome
        self.feasible = True
        self.fitness = self.callFitness()

    @classmethod
    def setProblem(cls, problem):
        cls.problem = problem

    def isFeasible(self):
        routes = self.chromosome.split(str(problem.depotCluster))
        for route in routes:
            routeDemand = 0
            for c in list(route.split()):
                nodes = problem.clusters[int(c)]
                for node in nodes:
                    routeDemand += node.demand
            if routeDemand > self.problem.capacity:
                return False
        return True

    @classmethod
    def createChromosome(cls):
        demands = []
        chromosome = ''

        for key in cls.problem.clusters:
            clust_demand = 0
            for node in cls.problem.clusters[key]:
                clust_demand += node.demand
            demands.append((key, clust_demand))

        depoClust = demands.pop(0)[0]
        amount = 0
        for _ in range(len(demands)):
            (cluster, demand) = rn.choice(demands)
            demands.remove((cluster, demand))

            amount += demand
            if(amount <= cls.problem.capacity):
                chromosome += str(cluster) + ' '
            else:
                chromosome += str(depoClust) + ' ' + str(cluster) + ' '
                amount = demand

        return chromosome

    def mutate(self):

        globalRoutes = self.chromosome.split(str(problem.depotCluster))

        while True:

            r1 = rn.randrange(len(globalRoutes))
            r2 = rn.randrange(len(globalRoutes))

            while r1 == r2:
                r2 = rn.randrange(len(globalRoutes))

            if len(globalRoutes[r1].split()) != 0 and len(globalRoutes[r2].split()) != 0:
                break

        n1 = rn.randrange(len(globalRoutes[r1].split()))
        n2 = rn.randrange(len(globalRoutes[r2].split()))

        temp = globalRoutes[r1].split()[n1]

        l1 = list(globalRoutes[r1].split())
        l1[n1] = globalRoutes[r2].split()[n2]

        l2 = list(globalRoutes[r2].split())
        l2[n2] = temp

        globalRoutes[r1] = ' '.join(l1)
        globalRoutes[r2] = ' '.join(l2)

        self.chromosome = ''
        for r in globalRoutes:
            for c in r.split():
                self.chromosome += c + ' '
            self.chromosome += str(problem.depotCluster) + ' '

        temp = self.chromosome.split()[:-1]
        self.chromosome = ' '.join([e for e in temp])

    def crossOver(self, parent2):
        size = len(self.chromosome.split())
        child1 = [-1 for i in range(size)]
        child2 = [-1 for i in range(size)]
        if XOVER_METHOD == ORDER_2POINT:
            ind2 = ind1 = rn.randrange(size)
            while ind2 == ind1:
                ind2 = rn.randrange(size)

            if ind2 < ind1:
                ind1, ind2 = ind2, ind1

            # transfer middle part to childs
            for i in range(ind1, ind2):
                child1[i] = self.chromosome.split()[i]
                child2[i] = parent2.chromosome.split()[i]

            # fill childs chromosome
            i1 = i2 = index = ind2
            depotCount = self.chromosome.count(str(self.problem.depotCluster))
            depot = str(self.problem.depotCluster)

            while (-1 in child1) or (-1 in child2):
                gen1 = self.chromosome.split()[index % size]

                gen2 = parent2.chromosome.split()[index % size]

                if (gen2 == depot and child1.count(depot) < depotCount) or not(gen2 in child1):
                    child1[i1 % size] = gen2
                    i1 += 1

                if (gen1 == depot and child2.count(depot) < depotCount) or not(gen1 in child2):
                    child2[i2 % size] = gen1
                    i2 += 1

                index += 1

        child1 = ' '.join([str(e) for e in child1])
        child2 = ' '.join([str(e) for e in child2])

        return Individual(child1), Individual(child2)

    def callFitness(self):
        if not self.isFeasible():
            return -1

        fitness = 0

        routes = self.chromosome.split(str(problem.depotCluster))

        depot = problem.clusters[problem.depotCluster][0]

        for route in routes:
            nodeList = [depot]

            for c in list(route.split()):

                nodes = problem.clusters[int(c)]

                for node in nodes:
                    nodeList.append(node)

            # now we have all nodes of one route
            # initial matrix of edges
            size = len(nodeList)
            mat = np.zeros(shape=(size, size))

            for i in range(size):
                for j in range(size):
                    if i != j:
                        n1 = nodeList[i]
                        n2 = nodeList[j]
                        distance = math.sqrt(
                            (n1.x - n2.x)**2 + (n1.y - n2.y)**2)

                        # nodes from the same cluster
                        if n1.cluster == n2.cluster:
                            mat[i][j] = distance

                        # adding M to distance of nodes from different clusters
                        else:
                            M = self.MList[n1.cluster][n2.cluster]
                            mat[i][j] = distance + M
                    else:
                        mat[i][j] = 0

            # now solving tsp of edges matrix
            solver = TSP()
            solver.read_mat(mat)

            sol = TwoOpt_solver(initial_tour='NN', iter_num=100)
            answer = solver.get_approx_solution(sol)

            cost = answer[0]
            nodes = answer[1]

            # decreasing total added Ms to answer
            for i in range(len(nodes)-1):
                n1 = nodeList[nodes[i]]
                n2 = nodeList[nodes[i+1]]

                if n1.cluster != n2.cluster:
                    cost -= self.MList[n1.cluster][n2.cluster]

            fitness += cost

        return fitness

    def __str__(self):
        # return self.chromosome[:10] + ' ...\t' +\
        #     'fitness: ' + str(self.fitness)
        return str(round(self.fitness, 3)) + ' , '+self.chromosome[0:20] + ' --- '

    def __repr__(self):
        return str(self)

    @classmethod
    def calculateMs(cls):
        clusterNum = len(cls.problem.clusters)
        mList = np.zeros(shape=(clusterNum, clusterNum))
        # print('clusters:', clusterNum)

        for i in range(clusterNum):
            m = 0
            for j in range(clusterNum):
                if i != j:
                    for n1 in cls.problem.clusters[i]:
                        for n2 in cls.problem.clusters[j]:
                            m += math.sqrt((n1.x - n2.x)**2 + (n1.y - n2.y)**2)

                    mList[i][j] = m
                else:
                    mList[i][j] = 0

        cls.MList = mList

    @classmethod
    def parentSelection(cls, population):
        parent1 = parent2 = None

        population = sorted(population, key=lambda x: x.fitness)
        index = int(0.3 * len(population))
        if SELECTION == RANDOM:
            while True:
                parent1 = rn.choice(list(population[0:index]))
                parent2 = rn.choice(list(population[0:index]))
                if len(parent1.chromosome) == len(parent2.chromosome):
                    break

        return parent1, parent2


def initialPop(problem):
    population = []
    Individual.setProblem(problem)
    Individual.calculateMs()

    for _ in range(POPULATION_SIZE):
        chrom = Individual.createChromosome()
        indiv = Individual(chrom)
        # print(indiv)
        population.append(indiv)

    return population


def GA(problem, initialPop, maxGeneration=100,
       mutation_rate=0.2, debug=True):

    generation = 1
    population = initialPop(problem)

    timer = time.time()
    if TIMER_MODE:
        maxGeneration = 100000000

    for _ in range(maxGeneration):

        if TIMER_MODE and time.time()-timer > 10:
            print('time out')
            return sorted(population, key=lambda x: x.fitness)[0]

        population = sorted(population, key=lambda x: x.fitness)

        best = population[0].fitness
        avg = np.mean([p.fitness for p in population])
        bests.append(best)
        averages.append(avg)

        if DEBUG:
            print("generation:", generation, " best: ", best, "avg: ", avg)

        new_generation = [population[0]]
        for _ in range(int(POPULATION_SIZE/2)-1):
            (parent1, parent2) = Individual.parentSelection(population)

            (child1, child2) = parent1.crossOver(parent2)

            if rn.random() < MUTATION_RATE and child1.isFeasible():
                child1.mutate()
            if rn.random() < MUTATION_RATE and child1.isFeasible():
                child2.mutate()

            child1.callFitness()
            child2.callFitness()

            if child1.fitness != -1:
                new_generation.append(child1)
            else:
                new_generation.append(parent1)

            if child2.fitness != -1:
                new_generation.append(child2)
            else:
                new_generation.append(parent2)

        new_generation = sorted(
            new_generation, reverse=True, key=lambda x: x.fitness)

        population = new_generation[:-1]
        generation += 1

    generation -= 1
    population = sorted(population, key=lambda x: x.fitness)

    # plotResult(generation, bests, averages)

    return population[0]


if __name__ == '__main__':

    myRow = 3

    for root, directories, filenames in os.walk("instances/GoldenWasilKellyAndChao_0.1/"):
        for filename in filenames:
            file = os.path.join(root, filename)
            problem = loadInstance(str(file))

            if TIMER_MODE:
                run = 1
            else:
                run = 10
    
            print('name: ', problem.name, ' dimention: ',
                  problem.dimention, ' capacity: ', problem.capacity)
            answers = []

            for _ in range(run):
                start = time.time()

                sol = GA(problem, initialPop, MAX_GENERATION,
                         MUTATION_RATE, DEBUG)

                vehicels = sol.chromosome.count(str(problem.depotCluster)) + 1

                duration = str(time.time() - start)[0:6]
                print('time: ', duration, 'fitness:',
                      round(sol.fitness, 2), 'vehicels: ', vehicels)

                answers.append(
                    (sol.chromosome, sol.fitness, duration, vehicels))

            printResult(answers)
            if EXEl_WRITE:
                writeResultToExel(filename, answers, myRow)
                myRow += 1
