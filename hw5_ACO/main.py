from pprint import pprint
import random as rn
import math
import copy
from random import randrange
import matplotlib.pyplot as plt
import time
import openpyxl
import os
import tsp

from tspy2 import TSP
from tspy2.solvers import NN_solver
from tspy2.solvers import TwoOpt_solver
import numpy as np
from collections import defaultdict

iterations = 10
ALPHA = RHO = 0.2
Beta = 2
Q0 = 0.9

DEBUG = False
TIMER_MODE = False
EXEl_WRITE = False


def euclideanDist_2D(a, b):
    return math.sqrt((a.x - b.x)**2 + (a.y - b.y)**2)


def rouletteWheel(arr):
    return 0


class Node(object):

    def __init__(self, number, x, y):
        self.number = number
        self.x = x
        self.y = y
        self.demand = None
        self.cluster = None

    def getCord(self):
        return (self.x, self.y)

    def __str__(self):
        return str(self.number) + str(self.getCord())

    def __repr__(self):
        return str(self)


class Problem(object):

    def __init__(self):
        self.name = None
        self.dimention = None
        self.capacity = None
        self.depotCluster = None
        self.clusters = []
        self.clusterDemands = []

    def calculateClusterDemands(self):

        self.clusterDemands = {}

        for key in self.clusters:
            clust_demand = 0
            for node in self.clusters[key]:
                clust_demand += node.demand
            # self.clusterDemands.append((key, clust_demand))
            self.clusterDemands[key] = clust_demand

        del self.clusterDemands[self.depotCluster]
        # print('demands: ', self.clusterDemands)

    def __str__(self):
        return 'depot cluster: ' + str(self.depotCluster)

    def __repr__(self):
        return str(self)


class Ant(object):

    def __init__(self, cluster):
        self.solution = [cluster]

    def __str__(self):
        return 'solution:' + str(self.solution)

    def __repr__(self):
        return str(self)


class Colony(object):

    def __init__(self, problem):
        self.ants = []
        self.best = None
        self.iterBests = []
        self.iterAvgs = []
        self.antsNum = max(math.ceil(len(problem.clusters) / 10), 2)
        self.problem = problem
        self.initialEta()
        # self.randPositioning()
        self.calculateMs()
        # self.calculateT0()
        self.T0 = 0.00005
        self.initialTrail()

    def calculateT0(self):

        demand = 0
        src = self.problem.depotCluster
        solution = ""

        for _ in range(len(self.problem.clusters)-1):
            candidates = self.minDists[src]
            candidates.pop(src)

            while True:
                dest = candidates.index(min(candidates))
                candidates[dest] = 1000000000000000
                if str(dest) not in solution:
                    break

            src = dest
            demand += self.problem.clusterDemands[dest]

            if demand > self.problem.capacity:
                demand = self.problem.clusterDemands[dest]
                solution += str(self.problem.depotCluster) + \
                    " " + str(dest) + " "
            else:
                solution += str(dest) + " "

        # print(self.problem.clusterDemands)
        # print(solution)
        # fit = self.costCalculate(solution)

        size = len(self.problem.clusters)
        size = 1
        self.T0 = 1/(size * self.costCalculate(solution))

    def calculateMs(self):
        clusterNum = len(self.problem.clusters)
        mList = np.zeros(shape=(clusterNum, clusterNum))
        # print('clusters:', clusterNum)

        for i in range(clusterNum):
            m = 0
            for j in range(clusterNum):
                if i != j:
                    for n1 in self.problem.clusters[i]:
                        for n2 in self.problem.clusters[j]:
                            m += math.sqrt((n1.x - n2.x)**2 + (n1.y - n2.y)**2)

                    mList[i][j] = m
                else:
                    mList[i][j] = 0

        self.Mlist = mList

    # calculating min distance of each per of clusters
    def initialEta(self):

        size = len(self.problem.clusters)

        # min distance between clusters
        self.minDists = [[0 for _ in range(size)]
                         for _ in range(size)]

        self.eta = defaultdict(dict)
        for i in range(len(self.problem.clusters)):
            for j in range(i+1, len(self.problem.clusters)):

                # distance not calculated
                if not(j in self.eta[i]):
                    cl1 = self.problem.clusters[i]
                    cl2 = self.problem.clusters[j]

                    minDist = 1000000000000
                    for n1 in cl1:
                        for n2 in cl2:
                            dist = euclideanDist_2D(n1, n2)
                            if minDist > dist:
                                minDist = dist

                    self.eta[i][j] = 1/minDist

                    self.minDists[i][j] = minDist
                    self.minDists[j][i] = minDist

        # print(self.eta)

    # initial pheromone matrix
    def initialTrail(self):
        size = len(self.problem.clusters)
        self.T = [[self.T0 for _ in range(size)] for _ in range(size)]

    def randPositioning(self):

        self.ants = []
        # remove depot cluster form cluster list
        clusters = list(self.problem.clusters.keys())
        clusters.remove(self.problem.depotCluster)

        # randomly position ants on n clusters
        for _ in range(self.antsNum):
            cluster = rn.choice(clusters)
            clusters.remove(cluster)
            self.ants.append(Ant(cluster))

    def antNextMove(self, antNum):

        antSolution = self.ants[antNum].solution

        # remove depot cluster form cluster list
        clusters = [c for c in list(
            self.problem.clusters.keys()) if c != self.problem.depotCluster]

        # current cluster of ant solution
        r = antSolution[-1]

        # destinatin cluster
        s = None

        # candidate clusters to move(not repetetive)
        candidates = [
            c for c in clusters if c not in antSolution]

        args = [self.T[r][u] *
                self.eta[min(r, u)][max(r, u)] ** Beta for u in candidates]

        q = rn.random()
        # exploitation
        if q < Q0:

            s = candidates[args.index(max(args))]

        # exploration
        else:
            p = [arg/sum(args) for arg in args]

            # roulette wheel
            s = candidates[rouletteWheel(p)]

        antSolution.append(s)

        # check vehicle capacity

        # all depot occurences
        depotIndices = list([idx for idx, val in enumerate(
            antSolution) if val == self.problem.depotCluster])

        lastindex = max(depotIndices) if depotIndices else -1

        demands = sum(self.problem.clusterDemands[c]
                      for c in (antSolution[lastindex+1:]))

        if demands > self.problem.capacity:

            # print('back to depot', antSolution)
            clust = antSolution.pop(-1)
            antSolution.append(self.problem.depotCluster)
            antSolution.append(clust)

        # else:
        #     print('under capacity')

        self.ants[antNum].solution = antSolution

        self.localPheromoneUpdate(r, s)

    def localPheromoneUpdate(self, r, s):

        self.T[r][s] = (1-RHO) * self.T[r][s] + RHO * self.T0
        self.T[s][r] = (1-RHO) * self.T[s][r] + RHO * self.T0

    def globalPheromoneUpdate(self):

        avg = 0

        # determining best solution
        solutions = [ant.solution for ant in self.ants]

        sol0 = " ".join(str(e) for e in solutions[0])
        best = (solutions[0], self.costCalculate(sol0))

        avg += best[1]

        for solution in solutions[1:]:
            fitness = self.costCalculate(" ".join(str(e) for e in solution))

            avg += fitness
            if fitness < best[1]:
                best = (solution, fitness)

        avg /= len(solutions)

        self.iterBests.append(best[1])
        self.iterAvgs.append(avg)

        if self.best == None or best[1] < self.best[1]:
            self.best = best

        # global pheromone update
        bestSol = self.best[0]

        for i in range(len(self.T)):
            for j in range(len(self.T)):
                deltaT = 0

                # if edge between i and j be in best solution(order of nodes is not important)
                # we update pheromone of edges (i,j) and (j,i)

                if i in bestSol and j in bestSol and + \
                        abs(bestSol.index(i) - bestSol.index(j)) == 1:

                    deltaT = 1/self.best[1]

                self.T[i][j] = (1-ALPHA) * self.T[i][j] + \
                    (ALPHA * deltaT)

                self.T[j][j] = (1-ALPHA) * self.T[j][i] + \
                    (ALPHA * deltaT)

    def costCalculate(self, solution):

        fitness = 0

        routes = solution.split(str(problem.depotCluster))

        depot = problem.clusters[problem.depotCluster][0]

        for route in routes:
            nodeList = [depot]

            for c in list(route.split()):

                nodes = problem.clusters[int(c)]

                for node in nodes:
                    nodeList.append(node)

            # now we have all nodes of one route
            # initial matrix of edges
            size = len(nodeList)
            mat = np.zeros(shape=(size, size))

            for i in range(size):
                for j in range(size):
                    if i != j:
                        n1 = nodeList[i]
                        n2 = nodeList[j]
                        distance = math.sqrt(
                            (n1.x - n2.x)**2 + (n1.y - n2.y)**2)

                        # nodes from the same cluster
                        if n1.cluster == n2.cluster:
                            mat[i][j] = distance

                        # adding M to distance of nodes from different clusters
                        else:
                            M = self.Mlist[n1.cluster][n2.cluster]
                            mat[i][j] = distance + M
                    else:
                        mat[i][j] = 0

            # now solving tsp of edges matrix
            solver = TSP()
            solver.read_mat(mat)

            sol = TwoOpt_solver(initial_tour='NN', iter_num=100)
            answer = solver.get_approx_solution(sol)

            cost = answer[0]
            nodes = answer[1]

            # decreasing total added Ms to answer
            for i in range(len(nodes)-1):
                n1 = nodeList[nodes[i]]
                n2 = nodeList[nodes[i+1]]

                if n1.cluster != n2.cluster:
                    cost -= self.Mlist[n1.cluster][n2.cluster]

            fitness += cost

        return fitness

    def __str__(self):
        return str(best) + 'as best answer by ' + str(len(ants)) + 'ants'

    def __repr__(self):
        return str(self)


def printResult(answers):

    minAns = min(answers, key=lambda t: t[1])
    maxAns = max(answers, key=lambda t: t[1])
    variance = round(math.sqrt(np.var([ans[1]for ans in answers])), 3)

    print("\nbest[0:10]=", minAns[0][0:10], "\tmin cost:", minAns[1])
    if TIMER_MODE == False:
        print("worst[0:10]=", maxAns[0][0:10], "\tmax cost:",
              max(answers, key=lambda t: t[1])[1])
        print("\naverage cost:", sum(ans[1] for ans in answers)/len(answers))
        print("\nvariance of costs:", variance)

        print("\nmin time:", min(answers, key=lambda t: t[2])[2])
        print("avg time:", str(sum(float(ans[2])
                                   for ans in answers)/len(answers))[0:6])
        print("max time:", max(answers, key=lambda t: t[2])[2])


def writeResultToExel(file_name, answers, myRow):
    minCost = min(answers, key=lambda t: t[1])[1]
    if TIMER_MODE == False:
        maxCost = max(answers, key=lambda t: t[1])[1]
        avgCost = sum(ans[1] for ans in answers)/len(answers)
        costVariance = round(math.sqrt(np.var([ans[1]for ans in answers])), 3)

        avgTime = str(sum(float(ans[2])for ans in answers)/len(answers))[0:6]

    wbkName = 'one_min_Results.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    for wks in wbk.worksheets:
        myCol = 7

        # wks.cell(row=myRow, column=1).value = file_name

        wks.cell(row=myRow, column=myCol).value = minCost
        if TIMER_MODE == False:
            wks.cell(row=myRow, column=myCol+1).value = avgCost
            wks.cell(row=myRow, column=myCol+2).value = maxCost
            wks.cell(row=myRow, column=myCol+3).value = costVariance

            wks.cell(row=myRow, column=myCol+4).value = avgTime

    wbk.save(wbkName)
    wbk.close


def plotResult(generation, bests, averages):

    plt.plot(list(range(generation)), bests, '-', color="gray",
             label='best of generations', linewidth=1.52)

    plt.xlabel('best solution')
    plt.ylabel('generation')

    plt.show()

    plt.plot(list(range(generation)), averages, 'bo-',
             label='avg of generations', linewidth=1.5)

    plt.xlabel('best solution')
    plt.ylabel('generation')

    plt.show()


def loadInstance(file):
    problem = Problem()
    with open(file) as f:
        lines = [line.rstrip() for line in f]

        problem.name = lines[0].split(':')[1]
        problem.dimention = int(lines[3].split(':')[1])
        problem.capacity = int(lines[4].split(':')[1])

        problem.clusters = {}

        # NODE_COORD_SECTION first index
        cordSecIndex = 8
        # DEMAND_SECTION first index
        demSecIndex = cordSecIndex + problem.dimention + 1
        # CLUSTER_SECTION first index
        cludSecIndex = demSecIndex + problem.dimention + 1

        # Golden type instance . its cluster number start from 1
        if problem.name == '':
            x = -1
        else:
            x = 0

        problem.depotCluster = int(lines[cludSecIndex].split()[1]) + x

        for i in range(problem.dimention):
            data = lines[cordSecIndex+i].split()
            node = Node(int(data[0]), float(data[1]), float(data[2]))

            node.demand = int(lines[demSecIndex+i].split()[1])
            node.cluster = int(lines[cludSecIndex+i].split()[1]) + x
            problem.clusters.setdefault(node.cluster, []).append(node)

    problem.calculateClusterDemands()
    return problem


def plotProgress(bests, avgs):

    plt.plot(list(range(len(bests))), bests, '-', color="gray",
             label='best of generations', linewidth=1.5)

    plt.xlabel('best solution')
    plt.ylabel('generation')

    plt.show()

    plt.plot(list(range(len(avgs))), avgs, 'bo-',
             label='avg of generations', linewidth=1.5)

    plt.xlabel('avg solution')
    plt.ylabel('generation')

    plt.show()


def ACS(problem, iterations=50, debug=True):

    timer = time.time()
    if TIMER_MODE:
        iterations = 10000000000

    # number of problem clusters to be seen
    size = len(problem.clusters) - 2

    # initial ants
    colony = Colony(problem)
    print('ants number: ', colony.antsNum)

    for _ in range(iterations):

        if TIMER_MODE and time.time()-timer > 60:
            print('time out')
            break

        colony.randPositioning()
        # solution completion condition
        for _ in range(size):

            # solution construction and local pheromone update
            for i in range(colony.antsNum):
                colony.antNextMove(i)

        # global pheromone update and set colony best solution
        colony.globalPheromoneUpdate()

        if DEBUG:
            print("best: ", colony.best[1], '\n')
            # print(colony.T[0][0], colony.T[0][2], '\n')

    # plotProgress(colony.iterBests,colony.iterAvgs)

    return colony.best


if __name__ == '__main__':

    myRow = 3

    for root, directories, filenames in os.walk("instances/GoldenWasilKellyAndChao_0.1/"):
        for filename in filenames:
            file = os.path.join(root, filename)
            problem = loadInstance(str(file))
            
            if TIMER_MODE:
                run = 1
            else:
                run = 10

            print('\nname: ', problem.name, ' dimention: ',
                  problem.dimention, ' capacity: ', problem.capacity)
            answers = []

            for _ in range(run):
                start = time.time()

                best = ACS(problem, iterations, DEBUG)

                duration = str(time.time() - start)[0:6]

                print('time: ', duration, '\tcost:',
                      round(best[1], 2), '\tsol:', best[0][:4])

                answers.append(
                    (best[0], best[1], duration))

            printResult(answers)

            if EXEl_WRITE:
                writeResultToExel(filename, answers, myRow)
            myRow += 1
